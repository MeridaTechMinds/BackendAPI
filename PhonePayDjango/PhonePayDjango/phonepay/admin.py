from django.contrib import admin
from .models import Details
# Register your models here.
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


class registrationAdmin(admin.ModelAdmin):
    list_display = ("name","email",'phone','amount')  
    actions = ['download_selected']
    def download_selected(self, request, queryset):
        # Create a new Excel workbook and add a worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        # Define the header
        header = ["name","email",'phone','amount']

        # Add header to the worksheet
        for col_num, header_text in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            cell = worksheet[f"{col_letter}1"]
            cell.value = header_text
            cell.alignment = Alignment(horizontal='center')

        # Add data to the worksheet
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field_name in enumerate(header, 1):
                col_letter = get_column_letter(col_num)
                cell = worksheet[f"{col_letter}{row_num}"]
                cell.value = getattr(obj, field_name, None)
                cell.alignment = Alignment(horizontal='center')

        # Create the HTTP response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="downloaded_data.xlsx"'
        workbook.save(response)

        return response


    download_selected.short_description = "Download selected items"

# Register the model with the custom admin class
admin.site.register(Details, registrationAdmin)
