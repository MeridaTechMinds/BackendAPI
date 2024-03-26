from django.contrib import admin
from django.http import HttpResponse
from .models import registration,Review
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
from .forms import *



class registrationAdmin(admin.ModelAdmin):
    list_display = ('created_at',"Name","Primary_contact",'Secondary_contact',"Location",'Email_Id','Soft_Skills','Technical_Skills','General_Skills','Current_CTC','Expected_CTC','Notice_Period','Current_designation','Applied_designation','experience','Job_portal_source','Contacted_by')  
    list_per_page = 10000
    search_fields = ('created_at',"Name","Primary_contact",'Secondary_contact',"Location",'Email_Id','Soft_Skills','Technical_Skills','General_Skills','Current_CTC','Expected_CTC','Notice_Period','Current_designation','Applied_designation','experience','Job_portal_source','Contacted_by')  
    list_filter = ('Primary_contact',) 
    
    actions = ['download_selected']


    def download_selected(self, request, queryset):
        # Create a new Excel workbook and add a worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Define the header
        header = ["Name", "Primary_contact", "Secondary_contact", "Location", "Email_Id", "Soft_Skills", "Technical_Skills", 
                "General_Skills", "Current_CTC", "Expected_CTC", "Notice_period", "Current_designation", 
                "Applied_designation", "experience", "Job_portal_source", "Contacted_by"]

        # Add header to the worksheet
        for col_num, header_text in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            cell = worksheet[f"{col_letter}1"]
            cell.value = header_text
            cell.alignment = Alignment(horizontal='center')

        # Add data to the worksheet
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field_name in enumerate(header, 1):
                col_letter = get_column_letter(col_num)
                cell = worksheet[f"{col_letter}{row_num}"]
                cell.value = getattr(obj, field_name, None)
                cell.alignment = Alignment(horizontal='center')

        # Create the HTTP response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        current_datetime = datetime.now()
        current_date = current_datetime.date()
        response['Content-Disposition'] = f'attachment; filename= "{current_date} .xlsx"'
        workbook.save(response)

        return response


    download_selected.short_description = "Download selected items"

    # change_list_template = 'admin/hr_app/change_list_results.html'

# Register the model with the custom admin class
admin.site.register(registration, registrationAdmin)
   
from django.contrib.auth.signals import user_logged_in


class ReviewAdmin(admin.ModelAdmin):
    list_display=('user_name',"ReviewedBy",'Moral_character','punctuality','health','behaviour','attitude','Career_goals','understanding_level',
    'possitive_attitude_and_mind','executive','responsibility','response_ability','team_handling','planing','communicate_ability',
    'passion','confidence','profissional_background','work_experience','knowledge_level','english_skils','other_languages',
    'consider_to_client','Internal_hiring','reject','created_at','remarks',
)
    list_per_page = 10000
    list_filter = ('Internal_hiring','reject',"consider_to_client") 
    autocomplete_fields = ['user_name'] 

    actions = ['download_selected']
    readonly_fields = ['ReviewedBy',]

    def save_model(self, request, obj, form, change):
        # Change the value of the field here
        obj.ReviewedBy = request.user.username
        super().save_model(request, obj, form, change)

        # it will give the exclude values in the onetoone field
    # def formfield_for_foreignkey(self, db_field, request, **kwargs):
    #     if db_field.name == 'user_name':
    #         # Get the queryset of registrations already associated with Review instances
    #         associated_registrations = Review.objects.exclude(user_name=None).values_list('user_name_id', flat=True)
    #         # Exclude these registrations from the choices available in the dropdown list
    #         kwargs["queryset"] = registration.objects.exclude(id__in=associated_registrations)
            
    #     return super().formfield_for_foreignkey(db_field, request, **kwargs)
        
      
    

    def download_selected(self, request, queryset):
        # Create a new Excel workbook and add a worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Define the header
        header = ['user_name.Name', 'user_name.Email_Id', 'user_name.Primary_contact', 'Moral_character',
                'punctuality', 'health', 'behaviour', 'attitude', 'Career_goals', 'understanding_level',
                'possitive_attitude_and_mind', 'executive', 'responsibility', 'response_ability', 'team_handling',
                'planing', 'communicate_ability', 'passion', 'confidence', 'profissional_background',
                'work_experience', 'knowledge_level', 'english_skils', 'other_languages',
                'consider_to_client', 'Internal_hiring', 'reject', 'created_at', 'remarks']

        # Add header to the worksheet
        for col_num, header_text in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            cell = worksheet[f"{col_letter}1"]
            cell.value = header_text
            cell.alignment = Alignment(horizontal='center')

        # Add data to the worksheet
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field_name in enumerate(header, 1):
                col_letter = get_column_letter(col_num)
                cell = worksheet[f"{col_letter}{row_num}"]
                
                # Check if the field name contains a dot (.), indicating a related field
                if '.' in field_name:
                    # Split the field name to get the related model and field
                    related_model_name, related_field_name = field_name.split('.')
                    # Get the related model instance
                    related_model_instance = getattr(obj, related_model_name)
                    # Get the value of the related field
                    field_value = getattr(related_model_instance, related_field_name, None)
                else:
                    # For non-related fields, directly get the value
                    field_value = getattr(obj, field_name, None)
                
                cell.value = field_value
                cell.alignment = Alignment(horizontal='center')

        # Create the HTTP response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        current_datetime = datetime.now()
        current_date = current_datetime.date()
        response['Content-Disposition'] = f'attachment; filename="{current_date}.xlsx"'
        workbook.save(response)

        return response

    download_selected.short_description = "Download selected items"
admin.site.register(Review, ReviewAdmin)
admin.site.site_header="Merida HR Adminstration"

