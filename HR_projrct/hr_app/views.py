from django.shortcuts import render,redirect

# Create your views here.

from django.shortcuts import render , HttpResponse
from .models import registration

from datetime import datetime
from django.core.mail import send_mail 
from django.conf import settings 
from django.contrib import messages
from datetime import date
from django.db.models import Q

from django.utils import timezone
from datetime import timedelta
from . models import Review


def register(request):
    if request.method =="POST":
        name = request.POST.get("name")
        pc = request.POST.get("pc")
        sc = request.POST.get("sc")
        location = request.POST.get("location")
        email = request.POST.get("mail")
        cctc = request.POST.get("cctc")
        ectc = request.POST.get("ectc")
        np = request.POST.get("np")
        d = request.POST.get("d")
        applied_d = request.POST.get("ad")
        exp = request.POST.get("exp")
        g_skills = request.POST.get("g_skills")
        t_skills = request.POST.get("t_skills")  
        s_skills = request.POST.get("s_skills")
        Jps = request.POST.get("htkau")
        c_by = request.POST.get("cb")
        
        created_at = datetime.now()
        # Use the default manager's create() method to create a new registration object
        reg_obj = registration.objects.create(
            Name=name,
            Primary_contact=pc,
            Secondary_contact=sc,
            Location=location,
            Email_Id=email,
            Current_CTC=cctc,
            Expected_CTC=ectc,
            Notice_Period=np,
            Current_designation=d,
            Applied_designation=applied_d,
            experience=exp,
            General_Skills=g_skills,
            Technical_Skills=t_skills,
            Soft_Skills=s_skills,
            Job_portal_source=Jps,
            Contacted_by=c_by,
            created_at=created_at
            )
        # print(settings.EMAIL_HOST_USER)
        subject='Interview Registration Details of '+name
        # send_mail(
        #     subject,
        #     f"Name: {name}\nPhone_number: {pc}\nEmail:{email}\n:Location: {location}\nGenaral Skills: {g_skills}\nTechnical Skills: {t_skills}\nSoft Skills: {s_skills}\nCurrent CTC: {cctc}\nNotice_Period: {np}\nCurrent_designation: {d}\nApplied_designation: {applied_d}\nExperience: {exp}\nJob Portal Source: {Jps}\nContacted By: {c_by}",
        #     settings.EMAIL_HOST_USER,
        #     ['businesshr@meridatechminds.com'],
        #     fail_silently=False,
        #     )
        return redirect("/success")
    return render(request,"form.html")


def applied_candidates(request):
    tc=registration.objects.all().count()
    
    today = date.today()
    candidates_Today= registration.objects.filter(created_date = today)
    Catc=registration.objects.filter(created_date=today).count()

    yesterday = timezone.now().date() - timedelta(days=30)
    candidates_yesterday = registration.objects.filter(created_date__range=(yesterday,today)).order_by('-created_at')
    Cayc=candidates_yesterday.count()
    
   
    consider_to_client_count=Review.objects.filter(consider_to_client="yes").count()
    internal_hiring_count=Review.objects.filter(Internal_hiring="yes").count()
    reject_count=Review.objects.filter(reject="yes").count()

    consider_to_client=Review.objects.filter(consider_to_client="yes").order_by('-created_at')
    internal_hiring=Review.objects.filter(Internal_hiring="yes").order_by('-created_at')
    reject=Review.objects.filter(reject="yes").order_by('-created_at')

    context={"consider_to_client_count":consider_to_client_count,
             "internal_hiring_count":internal_hiring_count,
             "reject_count":reject_count,
             "consider_to_client":consider_to_client,
             "internal_hiring":internal_hiring,
             "reject":reject,
             "CFT":candidates_Today,
             "CFY":candidates_yesterday,
             "CFTc":Catc,
             "CFYc":Cayc,
             "tc":tc,
             "candidates_Today":candidates_Today,
             "today_date":today,
             "yesterday":yesterday
            }
    if request.method=="POST":
        value=request.POST.get("search")
        date_value=request.POST.get("datesearch")
        if value:
            can_data=registration.objects.filter(Q(Name=value) | Q(Primary_contact=value))
            skills=registration.objects.filter(Q(Technical_Skills__icontains=value) |Q(General_Skills__icontains=value) |Q(Soft_Skills__icontains=value))
            
            return render(request,"filter.html",{"can_data":can_data,"skills":skills})
        elif date_value:
            SearchByDate=registration.objects.filter(Q(created_date=date_value))
            
            return render(request,"filter.html",{"date":SearchByDate})
    return render(request,"dashboard.html",context)

def review_filter(request,value):
    if value=="consider_to_client":
        consider_to_client=Review.objects.filter(consider_to_client="yes").order_by('-created_at')
        if request.method=="POST":
            value=request.POST.get("search")
            date_value=request.POST.get("datesearch")
            
            if value:
                
                can_data=Review.objects.filter(Q(user_name__Name=value) | Q(ReviewedBy=value)).order_by('-created_at')
                
                return render(request,"review_filter_search.html",{"can_data":can_data,"value":value})
        return render(request,"review_filter.html" ,{"consider_to_client":consider_to_client,"value":value})
    elif value=="internal_hiring":
        internal_hiring=Review.objects.filter(Internal_hiring="yes").order_by('-created_at')
        if request.method=="POST":
            value=request.POST.get("search")
            date_value=request.POST.get("datesearch")
           
            if value:
                
                can_data=Review.objects.filter(Q(user_name__Name=value) | Q(ReviewedBy=value)).order_by('-created_at')
               
                return render(request,"review_filter_search.html",{"can_data":can_data,"value":value})
        return render(request,"review_filter.html" ,{"internal_hiring":internal_hiring,"value":value})
    elif value=="reject":
        reject=Review.objects.filter(reject="yes")
        if request.method=="POST":
            value=request.POST.get("search")
            date_value=request.POST.get("datesearch")
            
            if value:
               
                can_data=Review.objects.filter(Q(user_name__Nname=value) | Q(ReviewedBy=value)).order_by('-created_at')
                
                return render(request,"review_filter_search.html",{"can_data":can_data,"value":value})
        return render(request,"review_filter.html" ,{"reject":reject,"value":value})
    return redirect("/dashboard")



def success(request):
    return render(request,"success.html")


def single_candidate(request,id):
    data=registration.objects.filter(id=id)
    return render(request,"single_can.html",{"data":data})

def today_candidates(request):
    today = date.today()
    candidates_Today=registration.objects.filter(created_date=today).order_by('-created_at')
    return render(request,"single_can.html",{"CFT":candidates_Today})
def month_candidates(request):
    today = date.today()
    yesterday = timezone.now().date() - timedelta(days=30)
    candidates_yesterday = registration.objects.filter(created_date__range=(yesterday,today)).order_by('-created_at')
    return render(request,"single_can.html",{"CFY":candidates_yesterday})

def total_candidates(request):
    data=registration.objects.all().order_by('-created_at')
    return render(request,"single_can.html",{"total_data":data})
 

from django.http import HttpResponse
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from django.urls import reverse

def download_selected(request,value):
    
    if value=="consider_to_client":
        queryset = Review.objects.filter(consider_to_client="yes")
        
    elif value=="internal_hiring":
        queryset = Review.objects.filter(Internal_hiring="yes")
    elif value=="reject":
        queryset = Review.objects.filter(reject="yes")
    else:
        queryset = Review.objects.filter(user_name__Name=value)  # Filter queryset as needed
    
    # Your existing code for generating Excel file
    # Make sure to replace 'user_name' with the actual field name if it's different
    
    # Create a new Excel workbook and add a worksheet
    if queryset.count()>0:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Define the header
        header = ["user_name.Name", 'user_name.Email_Id', 'user_name.Primary_contact', 'Moral_character',
                'punctuality', 'health', 'behaviour', 'attitude', 'Career_goals', 'understanding_level',
                'possitive_attitude_and_mind', 'executive', 'responsibility', 'response_ability', 'team_handling',
                'planing', 'communicate_ability', 'passion', 'confidence', 'profissional_background',
                'work_experience', 'knowledge_level', 'english_skils', 'other_languages',
                'consider_to_client', 'Internal_hiring', 'reject', 'created_at', 'remarks']

        # Add header to the worksheet
        for col_num, header_text in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            cell = worksheet[f"{col_letter}1"]
            cell.value = header_text
            cell.alignment = Alignment(horizontal='center')

        # Add data to the worksheet
        # if queryset is not None:
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field_name in enumerate(header, 1):
                col_letter = get_column_letter(col_num)
                cell = worksheet[f"{col_letter}{row_num}"]
                    
                    # Check if the field name contains a dot (.), indicating a related field
                if '.' in field_name:
                        # Split the field name to get the related model and field
                    related_model_name, related_field_name = field_name.split('.')
                        # Get the related model instance
                    related_model_instance = getattr(obj, related_model_name)
                        # Get the value of the related field
                    field_value = getattr(related_model_instance, related_field_name, None)
                else:
                        # For non-related fields, directly get the value
                    field_value = getattr(obj, field_name, None)
                    
                cell.value = field_value
                cell.alignment = Alignment(horizontal='center')

            # Create the HTTP response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        
        current_datetime = datetime.now()
        current_date = current_datetime.date()
        response['Content-Disposition'] = f'attachment; filename="{current_date}.xlsx"'
        workbook.save(response)    

        return response
    message="List is Empty You Can't Download"
    return redirect(f"{reverse('messages')}?message={message}")

def message_view(request):
    message = request.GET.get('message')
    return render(request,"messages.html",{"message":message})

from django.contrib.auth.forms import UserCreationForm ,AuthenticationForm
from .forms import signupform
from django.contrib.auth import authenticate ,login ,logout
def user_logout(request):
    if  request.user.is_authenticated:
        logout(request)
        messages.success(request,'Logout Successfully ! ')
    return redirect('/dashboard')

from django.shortcuts import render, get_object_or_404


def add_review(request):
    # instance = get_object_or_404(Review, pk=pk)
    # # Your logic for the review page
    # context = {'instance': instance.user_name}
     
    
    return render(request, 'hr_app/review_add.html')

# views.py











from django.shortcuts import render, redirect
from .forms import ReviewForm  # Import your ReviewForm

def add_review(request):
    if request.method == 'POST':
        form = ReviewForm(request.POST)
        if form.is_valid():
            # Save the form data to create a new review
            review = form.save()
            # Redirect to a success page, or you can redirect to the detail view of the created review
            return redirect('review_detail', pk=review.pk)  # Adjust the URL name as needed
    else:
        form = ReviewForm()
    return render(request, 'hr_app/add_review.html', {'form': form})


from rest_framework import viewsets
from .serializers import *
class RegistrationViewSet(viewsets.ModelViewSet):
    queryset = registration.objects.all()
    serializer_class = RegistrationSerializer